# Combined Invoice Pipeline Script

# ============================================================================
# INSTALLATION COMMANDS (uncomment if needed)
# ============================================================================
# !pip install frontend
# !pip install pymupdf
# !pip install opencv-python
# !pip install pytesseract
# !pip install "python-doctr[torch]"

import os
import fitz  # PyMuPDF
import numpy as np
from PIL import Image
import io
from doctr.io import DocumentFile
from doctr.models import ocr_predictor

# ============================================================================
# NETWORK PATH CONFIGURATION
# ============================================================================
# Import network configuration from external file
try:
    import sys
    import os
    # Add the parent directory to the Python path to find network_config.py
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from network_config import NETWORK_CONFIG, get_network_path, get_local_fallback, is_network_enabled
    print("SUCCESS: Loaded network configuration from network_config.py")
except ImportError:
    print("WARNING: network_config.py not found, using default configuration")
    # Default configuration if external file not found
    NETWORK_CONFIG = {
        "projects_server": "192.168.1.130",
        "projects_share": "Projects", 
        "projects_folder": "Raj",
        "enable_network": True,
        "local_fallback": "data/network_fallback/"
    }
    
    def get_network_path():
        return f"\\\\{NETWORK_CONFIG['projects_server']}\\{NETWORK_CONFIG['projects_share']}\\{NETWORK_CONFIG['projects_folder']}\\"
    
    def get_local_fallback():
        return NETWORK_CONFIG["local_fallback"]
    
    def is_network_enabled():
        return NETWORK_CONFIG.get("enable_network", True)

# Configure network paths
NETWORK_PATHS = {
    "projects_folder": get_network_path(),
    "local_fallback": get_local_fallback()
}

def get_accessible_path(path_key):
    """
    Try network path first, then fallback to local path.
    Returns the first accessible path or None if neither works.
    """
    network_path = NETWORK_PATHS.get(path_key)
    local_path = NETWORK_PATHS.get("local_fallback")
    
    # Check if network is enabled
    if not is_network_enabled():
        print(f"WARNING: Network disabled, using local fallback: {local_path}")
        if local_path:
            os.makedirs(local_path, exist_ok=True)
            return local_path
        return None
    
    # Try network path first
    if network_path and os.path.exists(network_path):
        print(f"SUCCESS: Using network path: {network_path}")
        return network_path
    
    # Fallback to local path
    if local_path:
        os.makedirs(local_path, exist_ok=True)
        print(f"WARNING: Network path not accessible, using local fallback: {local_path}")
        return local_path
    
    print(f"ERROR: Neither network nor local path accessible for: {path_key}")
    return None

# Test network connectivity
PROJECTS_PATH = get_accessible_path("projects_folder")
if not PROJECTS_PATH:
    print("WARNING: Network path not accessible. Some features may be limited.")
    print("   To enable full functionality, ensure network path is accessible or")
    print("   copy required files to the local fallback folder.")
    
    # Create sample files in local fallback for testing
    def create_sample_files():
        """Create sample files in local fallback directory for testing"""
        fallback_dir = NETWORK_PATHS["local_fallback"]
        os.makedirs(fallback_dir, exist_ok=True)
        
        # Create sample PDF files with PO/Job numbers
        sample_files = [
            "PO_12345_sample.pdf",
            "Job_25.16_sample.pdf", 
            "PO_67890_sample.pdf",
            "Job_25.02_sample.pdf"
        ]
        
        for filename in sample_files:
            filepath = os.path.join(fallback_dir, filename)
            if not os.path.exists(filepath):
                # Create empty file for testing
                with open(filepath, 'w') as f:
                    f.write(f"Sample file: {filename}")
                print(f"Created sample file: {filepath}")
    
    create_sample_files()

# Load OCR model (CPU or GPU)
model = ocr_predictor(pretrained=True)

def pdf_to_text_doctr(pdf_path):
    """Extract text from PDF using DocTR."""
    text_output = ""
    try:
        # Load PDF as images
        doc_images = DocumentFile.from_pdf(pdf_path)

        # Perform OCR
        result = model(doc_images)

        # Extract text from each page
        for i, page in enumerate(result.pages):
            page_text = ""
            for block in page.blocks:
                for line in block.lines:
                    line_text = " ".join([word.value for word in line.words])
                    page_text += line_text + "\n"

            text_output += f"\n--- Page {i+1} ---\n{page_text}"

    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
    return text_output

def process_pdf_folder_doctr(folder_path, output_folder):
    """Processes all PDFs using DocTR and saves OCR output to text files."""
    os.makedirs(output_folder, exist_ok=True)

    for file in os.listdir(folder_path):
        if file.lower().endswith('.pdf'):
            file_path = os.path.join(folder_path, file)
            print(f"Processing: {file}")

            ocr_text = pdf_to_text_doctr(file_path)

            output_path = os.path.join(output_folder, f"{os.path.splitext(file)[0]}.txt")
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(ocr_text)

            print(f"Saved OCR text to: {output_path}")

# Example usage
input_pdf_folder = "data/raw_pdfs"              
output_text_folder = "data/OCR_text_Test"      
process_pdf_folder_doctr(input_pdf_folder, output_text_folder)

# !pip install -U bitsandbytes
# !pip install -U accelerate
# !pip install transformers
# !pip install sentencepiece

import requests

# Recommended: store your API key securely
MISTRAL_API_KEY = "5GPPqcV6edATEGnl0w09ORmhu8zqIzUL"

def call_mistral_api(prompt):
    url = "https://api.mistral.ai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {MISTRAL_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "pixtral-12b-2409",
        "messages": [
            {"role": "system", "content": "You are an intelligent invoice extraction system."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.1,
        "max_tokens": 1024
    }

    response = requests.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return response.json()["choices"][0]["message"]["content"]
    else:
        raise Exception(f"ERROR: API Error {response.status_code}: {response.text}")


def clean_ocr_text(text: str) -> str:
    # Remove placeholder page markers and empty quotes
    text = re.sub(r"'''\s*--- Page \d+ ---\s*'z''", "", text)
    text = re.sub(r"--- Page \d+ ---", "", text)
    return text.strip()

import re

def build_core_prompt(ocr_text: str) -> str:
    return f"""
You are an expert invoice data extraction system designed for messy OCR-scanned documents from various vendors. Extract exactly the 7 fields below from the OCR text and return a flat, clean **JSON object**.

If any field is unclear or missing, return it as an empty string `""`. Do not guess. Do not add extra fields or formatting.

---

📌 FIELDS TO EXTRACT:

1. **GL_Date**
   - Format: MM/DD/YY
   - Use Invoice_Date if GL_Date is missing (they are always the same).

2. **Invoice_Number**
   - Look for: "Invoice #", "Invoice No.", "Inv #", "Doc #", "Credit Memo", or any line that contains "Credit" followed by a number
   - Extract only the number — do not include words like "Credit Memo"
   - Examples: From "CREDIT SLAKEY MEMO 859388323", extract "859388323"
   - ⛔ Avoid: "Customer Number", "Cust No.", "Account #"
   - Remove symbols like `#`, `:`.

3. **Invoice_Date**
   - Look for: "Invoice Date", "Date Issued", "Inv Date"
   - Format: MM/DD/YY
   - If multiple dates found, return the earliest valid one.

4. **Invoice_Amount**
   - Look for: "Total Due", "Invoice Amount", "Amount Due", "Balance Due"
   - ⛔ Avoid: "Subtotal", "Tax", "Discount", "Previous Balance"
   - Extract the **largest visible amount** from valid labels
   - CRITICAL: Preserve decimal points exactly as they appear (e.g., "48529.29" should remain "48529.29", not "4852929")
   - Remove symbols: `$`, `,`, `%`, `(`, `)`
   - Only treat the amount as negative if a clear `-` appears directly in front of the number — parentheses alone do not imply negative
   - Example: If you see "$48,529.29", extract "48529.29"

5. **Amount_Before_Taxes**
   - Look for: "Subtotal", "Amount Before Tax", "Total Before Tax"
   - ⛔ Avoid: "Tax", "Total Due", "Invoice Total", "Balance Due"
   - CRITICAL: Preserve decimal points exactly as they appear
   - Remove symbols: `$`, `,`, `%`, `(`, `)`
   - Only treat the amount as negative if a clear `-` appears directly in front of the number — parentheses alone do not imply negative
   - Example: If you see "$45,000.00", extract "45000.00"

6. **Tax_Amount**
   - Look for: "Sales Tax", "Tax", "Tax Amount", "VAT"
   - ⛔ Avoid: "Total", "Subtotal", "Before Tax", "After Tax"
   - Extract the first valid amount shown near the correct label — even if negative
   - CRITICAL: Preserve decimal points exactly as they appear
   - Include negative sign if a clear `-` appears directly in front of the number
   - Remove all symbols: `$`, `,`, `%`, `(`, `)`
   - Only treat the amount as negative if a clear `-` appears directly in front of the number — parentheses alone do not imply negative
   - Example: If you see "$3,529.29", extract "3529.29"

7. **Shipping_Charges**
   - Look for: "Shipping", "Freight", "Delivery", "Handling", "S&H", "Shipping & Handling"
   - Extract the amount shown next to these labels
   - CRITICAL: Preserve decimal points exactly as they appear
   - Remove symbols: `$`, `,`, `%`, `(`, `)`
   - Only treat as negative if a clear `-` appears directly in front of the number
   - If no shipping charges found, return empty string
   - Example: If you see "$25.50", extract "25.50"

---

❗ LOGICAL RULES:
- If Tax_Amount is present and greater than 0, then Invoice_Amount and Amount_Before_Taxes **must be different**, based on the actual visible values
- Never extract the same amount for both fields if Tax_Amount is present
- Only use amounts shown clearly next to their correct labels — no guessing or assumptions
- CRITICAL: Never hallucinate amounts that are not clearly visible in the OCR text

---

📏 OUTPUT RULES:
- Dates must be in MM/DD/YY format
- Amounts are clean numbers without formatting but WITH decimal points preserved
- Parentheses do not imply negative unless a clear `-` is present
- Return a clean JSON with exactly the 7 fields — no extra text or formatting
- CRITICAL: Only extract amounts that are explicitly visible in the OCR text

---

🧾 OCR TEXT START:
{ocr_text}
🧾 OCR TEXT END

---

SUCCESS: RETURN:
A JSON object with exactly the 7 fields. Leave missing fields as empty strings `""`.
"""

import re

def clean_json_output(text: str) -> str:
    """
    Cleans model output containing JSON by:
    - Removing trailing commas
    - Stripping ```json or ``` markdown blocks
    - Stripping whitespace
    """
    cleaned = re.sub(r",\s*([}\]])", r"\1", text)  # remove trailing commas
    cleaned = re.sub(r"```(?:json)?", "", cleaned)  # remove ```json or ```
    cleaned = cleaned.strip()
    return cleaned

import re

def strip_labels(value: str) -> str:
    """
    Remove label fragments like 'Credit', 'Memo', 'Invoice', etc. from Invoice_Number field.
    """
    return re.sub(r'(?i)(invoice|credit|memo|number|#|doc|no\.?)', '', value).strip()

import re

def clean_amount(val):
    if val == "" or val is None:
        return None
    val = str(val)  # Ensure it's a string before .replace()
    # Remove currency symbols and formatting, but preserve decimal points
    cleaned = val.replace("$", "").replace(",", "").replace("(", "").replace(")", "").replace("%", "").replace("USD", "").strip()
    
    try:
        # Convert to float and format with 2 decimal places
        float_val = float(cleaned)
        return f"{float_val:.2f}"
    except ValueError:
        return None

def amount_exists_in_text(amount, ocr_text):
    """Check if a specific amount exists in OCR text with 2 decimal formatting"""
    pattern = rf'{amount:,.2f}'.replace(",", r"[,]*")
    return bool(re.search(pattern, ocr_text))

def correct_amounts(result_json, ocr_text):
    A = clean_amount(result_json.get("Invoice_Amount", ""))
    B = clean_amount(result_json.get("Amount_Before_Taxes", ""))
    C = clean_amount(result_json.get("Tax_Amount", ""))

    ocr_amounts = []
    for amt in re.findall(r'[\d,]+\.\d{2}', ocr_text):
        try:
            ocr_amounts.append(round(float(amt.replace(",", "")), 2))
        except:
            continue
    ocr_amounts = set(ocr_amounts)

    # Convert string amounts to floats for numerical comparisons
    A_float = float(A) if A and A.replace('.', '').replace('-', '').isdigit() else None
    B_float = float(B) if B and B.replace('.', '').replace('-', '').isdigit() else None
    C_float = float(C) if C and C.replace('.', '').replace('-', '').isdigit() else None

    # Fix: If Tax_Amount present & Invoice_Amount == Amount_Before_Taxes
    if C_float and C_float > 0 and A_float == B_float and A_float is not None:
        estimated_B = round(A_float - C_float, 2)
        if estimated_B in ocr_amounts:
            B = str(estimated_B)
            result_json["Amount_Before_Taxes"] = B
        else:
            result_json["Amount_Before_Taxes"] = ""

    # Recover missing amounts
    if A_float is None and B_float is not None and C_float is not None:
        estimated_A = round(B_float + C_float, 2)
        if estimated_A in ocr_amounts:
            A = str(estimated_A)
            result_json["Invoice_Amount"] = A

    if B_float is None and A_float is not None and C_float is not None:
        estimated_B = round(A_float - C_float, 2)
        if estimated_B in ocr_amounts:
            B = str(estimated_B)
            result_json["Amount_Before_Taxes"] = B

    # Final Logical Check
    if A_float and B_float and C_float:
        if not (A_float > B_float and A_float > C_float and B_float >= 0 and C_float >= 0):
            result_json["Invoice_Amount"] = A if A else ""
            result_json["Amount_Before_Taxes"] = B if B and A_float > B_float else ""
            result_json["Tax_Amount"] = C if C and A_float > C_float else ""

    return result_json

import re

def fallback_extract_amounts(ocr_text):
    """
    Extract Subtotal and Tax directly from OCR using regex,
    used only when the model fails to do so.
    """
    subtotal_match = re.search(r"Sub[-\s]?Total\s*[:\-]?\s*(-?\$?\d[\d,\.]*)", ocr_text, re.IGNORECASE)
    tax_match = re.search(r"\bTax\s*[:\-]?\s*(-?\$?\d[\d,\.]*)", ocr_text, re.IGNORECASE)

    subtotal = subtotal_match.group(1).replace("$", "").replace(",", "") if subtotal_match else ""
    tax = tax_match.group(1).replace("$", "").replace(",", "") if tax_match else ""

    return subtotal.strip(), tax.strip()

def fix_truncated_json(json_str: str) -> str:
    """
    Tries to fix a truncated or incomplete JSON string by:
    - Ensuring closing brackets are added if missing
    - Truncating after the last closing brace
    """
    # Remove trailing commas
    json_str = re.sub(r",\s*([}\]])", r"\1", json_str)

    # Ensure JSON ends with closing brace
    last_open = json_str.rfind("{")
    last_close = json_str.rfind("}")
    if last_open != -1 and last_close == -1:
        json_str += "\n\"AP_GL_Account\": \"\",\n\"Distribution_GL_Account\": \"\",\n\"Job_Number\": \"\"\n}"

    # Clean up excess or nested block
    if json_str.count("{") > json_str.count("}"):
        json_str += "}"

    return json_str.strip()

import json
import re

def extract_json_from_response(response: str):
    """
    Extracts and parses a clean JSON object from a language model's raw response string.
    """
    try:
        # Step 1: Clean up formatting artifacts
        response = clean_json_output(response)

        # Step 2: Extract JSON block
        start = response.find("{")
        end = response.rfind("}")
        if start == -1 or end == -1 or end <= start:
            return None, "No valid JSON braces found"

        json_str = response[start:end + 1]

        # Step 3: Parse the JSON
        parsed = json.loads(json_str)
        return parsed, None

    except Exception as e:
        return None, f"JSON Parsing Error: {str(e)}"

import os
import json
import re

ocr_dir = "data/OCR_text_Test"
output_dir = "data/processed"
failed_log_path = os.path.join(output_dir, "failed_log.jsonl")
os.makedirs(output_dir, exist_ok=True)

for filename in os.listdir(ocr_dir):
    if filename.endswith(".txt"):
        with open(os.path.join(ocr_dir, filename), "r", encoding="utf-8") as f:
            raw_ocr = f.read()

        ocr_text = clean_ocr_text(raw_ocr)
        prompt = build_core_prompt(ocr_text)
        result = call_mistral_api(prompt)

        parsed_data, error_reason = extract_json_from_response(result)

        if parsed_data:
            parsed_data["Invoice_Number"] = strip_labels(parsed_data.get("Invoice_Number", ""))
            parsed_data = correct_amounts(parsed_data, ocr_text)

            # Fallback if model fails to extract subtotal or tax
            if not parsed_data.get("Amount_Before_Taxes"):
                fallback_subtotal, _ = fallback_extract_amounts(ocr_text)
                parsed_data["Amount_Before_Taxes"] = fallback_subtotal
        
            if not parsed_data.get("Tax_Amount"):
                _, fallback_tax = fallback_extract_amounts(ocr_text)
                parsed_data["Tax_Amount"] = fallback_tax
            
            # Ensure Shipping_Charges field exists
            if "Shipping_Charges" not in parsed_data:
                parsed_data["Shipping_Charges"] = ""

            parsed_data["__prompt"] = prompt  # Optionally store for trace
            out_file = filename.replace(".txt", ".json")
            with open(os.path.join(output_dir, out_file), "w", encoding="utf-8") as f:
                json.dump(parsed_data, f, indent=2)
            print(f"Extracted: {filename}")
        else:
            print(f"Failed to extract from: {filename}")
            with open(failed_log_path, "a", encoding="utf-8") as log_f:
                json.dump({
                    "file": filename,
                    "reason": error_reason,
                    "raw_output": result.strip()
                }, log_f)
                log_f.write("\n")

import os
import re
import json
import pandas as pd
from fuzzywuzzy import fuzz
import time

# --- Load and prepare vendor list ---
vendor_df = pd.read_csv("data/Vendor_List.csv").dropna(subset=['Vendor_Name'])
vendor_df['Vendor_Contact'] = vendor_df['Vendor_Contact'].fillna("")
vendor_df['Vendor_Address'] = vendor_df['Vendor_Address'].fillna("")
vendor_df["Vendor_Code"] = vendor_df["Vendor_Code"].astype(str).str.strip()

# --- Prepare OCR and JSON folders ---
ocr_txt_folder = "data/OCR_text_Test"
json_folder = "data/processed"
output_csv = "outputs/excel_files/matched_vendors_from_txt.csv"
approval_flag_file = "outputs/excel_files/approval_needed.flag"
approval_status_file = "outputs/excel_files/approval_status.json"

# Clear any existing approval files at the start of processing
if os.path.exists(approval_flag_file):
    os.remove(approval_flag_file)
    print("SUCCESS: Cleared existing approval flag")

if os.path.exists(approval_status_file):
    os.remove(approval_status_file)
    print("SUCCESS: Cleared existing approval status")

# Clear any existing PO approval files at the start of processing
po_approval_flag_file = "outputs/excel_files/po_approval_needed.flag"
if os.path.exists(po_approval_flag_file):
    os.remove(po_approval_flag_file)
    print("SUCCESS: Cleared existing PO approval flag")

# --- Utility Functions ---
def normalize_digits(text):
    return re.sub(r'\D+', '', str(text))

def extract_contact_numbers(text):
    return re.findall(r'\b\d{10,}\b', normalize_digits(text))

def remove_company_address(text):
    for block in ["535 railroad ave", "535 railroad avenue"]:
        text = text.replace(block, "")
    return text

def wait_for_approval():
    """Wait for user approval before continuing"""
    print("SUCCESS: Vendor matches created. Waiting for user approval...")
    
    # Create approval flag
    with open(approval_flag_file, 'w') as f:
        f.write("approval_needed")
    
    # Initialize approval status
    approval_status = {"approved": False, "approved_matches": []}
    with open(approval_status_file, 'w') as f:
        json.dump(approval_status, f)
    
    # Wait for approval
    while True:
        try:
            with open(approval_status_file, 'r') as f:
                status = json.load(f)
            
            if status.get("approved", False):
                print("SUCCESS: Approval received. Continuing with processing...")
                # Remove approval flag
                if os.path.exists(approval_flag_file):
                    os.remove(approval_flag_file)
                return status.get("approved_matches", [])
            
            time.sleep(1)  # Check every second
            
        except Exception as e:
            print(f"Error checking approval status: {e}")
            time.sleep(1)

# --- Step 1: Match Vendors from .txt using address/contact ---
final_matches = []

for txt_file in os.listdir(ocr_txt_folder):
    if not txt_file.endswith(".txt"):
        continue

    with open(os.path.join(ocr_txt_folder, txt_file), 'r', encoding='utf-8') as f:
        txt_content = f.read()
        normalized_txt = remove_company_address(txt_content.lower())
        txt_digits = extract_contact_numbers(txt_content)

        best_match = None
        best_score = 0

        for _, row in vendor_df.iterrows():
            contact = normalize_digits(row['Vendor_Contact'])
            address = str(row['Vendor_Address']).lower()

            contact_matched = any(contact in c for c in txt_digits) if contact else False
            address_score = fuzz.token_set_ratio(address, normalized_txt) if address else 0
            address_matched = address_score >= 80

            if address_matched or contact_matched:
                combined_score = address_score + (10 if contact_matched else 0)

                if combined_score > best_score:
                    best_score = combined_score
                    best_match = {
                        "TXT_File": txt_file,
                        "Vendor_Code": row['Vendor_Code'],
                        "Vendor_Name": row['Vendor_Name'],
                        "Vendor_Contact": row['Vendor_Contact'],
                        "Vendor_Address": row['Vendor_Address'],
                        "Address_Match_Score": address_score,
                        "Matched_Contact": contact if contact_matched else "",
                        "Matched_By": "contact + address" if (contact_matched and address_matched)
                                      else "address only" if address_matched else "contact only"
                    }

        if best_match:
            final_matches.append(best_match)

# --- Step 2: Save Matches to CSV ---
match_df = pd.DataFrame(final_matches)

# Clean up addresses to replace newlines with spaces before writing CSV
if 'Vendor_Address' in match_df.columns:
    match_df['Vendor_Address'] = match_df['Vendor_Address'].astype(str).str.replace('\n', ' ').str.replace('\r', ' ')

# Write CSV with proper quoting to handle any remaining special characters
match_df.to_csv(output_csv, index=False, quoting=1)  # quoting=1 means quote all non-numeric fields
print(f"SUCCESS: Best vendor matches saved to {output_csv}")

# --- Step 3: Wait for User Approval ---
approved_matches = wait_for_approval()

# --- Step 4: Continue with approved matches only ---
if approved_matches:
    # Filter matches to only include approved ones
    approved_txt_files = [match["TXT_File"] for match in approved_matches]
    final_matches = [match for match in final_matches if match["TXT_File"] in approved_txt_files]
    
    print(f"SUCCESS: Processing {len(final_matches)} approved vendor matches...")
    
    # IMPORTANT: Re-read the CSV file to get the updated vendor information
    print("SUCCESS: Reading updated vendor information from CSV file...")
    updated_match_df = pd.read_csv(output_csv)
    print(f"SUCCESS: Loaded {len(updated_match_df)} updated vendor matches from CSV")
    
else:
    print("WARNING: No matches approved. Stopping processing.")
    exit(0)

# --- Step 5: Prepare Vendor Details Lookup ---
vendor_lookup = {
    str(row["Vendor_Code"]).strip(): {
        "Distribution_GL_Account": row.get("Distribution_GL_Account", ""),
        "Phase_Code": row.get("Phase_Code", ""),
        "Cost_Type": row.get("Cost_Type", "")
    }
    for _, row in vendor_df.iterrows()
}

# --- Step 6: Inject only approved vendor info into JSON files ---
for _, row in updated_match_df.iterrows():
    if row["TXT_File"] not in approved_txt_files:
        continue
        
    base_name = os.path.splitext(row["TXT_File"])[0]
    json_path = os.path.join(json_folder, base_name + ".json")

    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Inject matched vendor info using the UPDATED data from CSV
        data["Vendor_Code"] = row["Vendor_Code"]
        data["Vendor_Name"] = row["Vendor_Name"]

        # ERROR: Remove logic for Distribution_GL_Account, Phase_Code, Cost_Type

        # Save updated JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        print(f"SUCCESS: Enriched {json_path} with updated vendor info: {row['Vendor_Code']} - {row['Vendor_Name']}")
    else:
        print(f"ERROR: JSON not found for: {base_name}")

print("SUCCESS: Processing complete!")

import os
os.environ["MISTRAL_API_KEY"] = "5GPPqcV6edATEGnl0w09ORmhu8zqIzUL"  # Replace with your actual key

import os
import base64
import pandas as pd
from pdf2image import convert_from_path
from mistralai import Mistral
from dotenv import load_dotenv
import re

load_dotenv()
os.environ["TOKENIZERS_PARALLELISM"] = "false"

# Get the absolute path to the poppler bin directory
import os
current_dir = os.path.dirname(os.path.abspath(__file__))
POPPLER_PATH = os.path.join(current_dir, "..", "models", "poppler-24.08.0", "Library", "bin")

def convert_pdf_to_image(pdf_path, page_number=0, poppler_path=POPPLER_PATH):
    try:
        # Debug: Check if poppler path exists
        if not os.path.exists(poppler_path):
            print(f"WARNING: Poppler path does not exist: {poppler_path}")
            return None
            
        # Debug: Check if pdfinfo exists in the poppler path
        pdfinfo_path = os.path.join(poppler_path, "pdfinfo.exe")
        if not os.path.exists(pdfinfo_path):
            print(f"WARNING: pdfinfo not found at: {pdfinfo_path}")
            return None
            
        print(f"SEARCH: Using poppler path: {poppler_path}")
        images = convert_from_path(pdf_path, dpi=200, poppler_path=poppler_path)
        if not images:
            raise ValueError("No images generated from PDF.")
        image_path = f"{os.path.splitext(os.path.basename(pdf_path))[0]}_page.jpg"
        images[page_number].save(image_path, "JPEG")
        return image_path
    except Exception as e:
        print(f"Error converting {pdf_path} to image: {e}")
        return None

def encode_image(image_path):
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        print(f"Error encoding image: {e}")
        return None

def extract_po_from_image(base64_image, client, model="pixtral-12b-2409"):
    try:
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": (
                        "Extract the Customer PO number from this invoice EXACTLY as it appears, preserving all formatting, punctuation, decimal points, spaces, and dashes. "
                        "PO numbers can be: short phrases (e.g., 'SHOP', 'ELECTRIC WALL HTR'), numbers with decimals (e.g., '24.08', '22.82'), or plain numbers (e.g., '1234', '56789'). "
                        "CRITICAL: Do NOT modify, combine, or change the format. If you see '24.08', return '24.08' NOT '2408'. "
                        "Ignore invoice numbers, order numbers, totals, and dates. "
                        "Only return the PO number exactly as written — do not guess, hallucinate, or reformat. If not visible, return 'NOT FOUND'."
                    )},
                    {"type": "image_url", "image_url": f"data:image/jpeg;base64,{base64_image}"}
                ]
            }
        ]
        response = client.chat.complete(model=model, messages=messages)
        result = response.choices[0].message.content.strip()
        return "" if result.strip().upper() == "NOT FOUND" else result
    except Exception as e:
        return f"ERROR: {e}"

def extract_amount_before_tax(text):
    patterns = [r"(?:subtotal|amount\s*before\s*tax|total\s*before\s*tax)[:\s]*\$?\s*([0-9,]+\.\d{2})"]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).replace(',', '')
    return ""

def is_valid_po(candidate: str) -> bool:
    COMMON_PO_SKIP_WORDS = {"description", "amount", "invoice", "sales", "total", "terms", "date"}
    candidate = candidate.strip().upper()
    if candidate.lower() in COMMON_PO_SKIP_WORDS:
        return False
    if re.fullmatch(r"\d{4,6}", candidate):
        return True
    if re.fullmatch(r"\d{2}\.\d{2,3}", candidate):
        return True
    if re.fullmatch(r"[A-Z0-9\- ]{3,}", candidate):
        return True
    return False

def clean_po_value(raw_text):
    if not raw_text:
        return ""
    text = raw_text.strip().upper()
    
    # Handle PO # pattern (with space) - extract just the number
    po_hash_match = re.search(r"PO\s+#(\d+)", text, re.IGNORECASE)
    if po_hash_match:
        return po_hash_match.group(1)
    
    # Updated cleaning logic (safe)
    text = re.sub(r"\bPO\s*NUMBER\s*[:\-]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCUSTOMER\s*PO\s*[:\-]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bPURCHASE\s*ORDER\s*[:\-]?\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCUSTOMER\s*ORDER\s*NUMBER\s*[:\-]?\s*", "", text, flags=re.IGNORECASE)

    tokens = re.split(r"[\n,:;\-]+", text)
    for token in tokens:
        token = token.strip()
        if is_valid_po(token):
            return token
    if is_valid_po(text):
        return text
    return ""


def classify_po(value):
    try:
        val = str(value).strip()

        # Case 1: 4-digit number → PO_Number
        if re.fullmatch(r'\d{4}', val):
            return val, '', '', ''

        # Case 2: 5-digit number → WO_Number
        if re.fullmatch(r'\d{5}', val):
            return '', '', val, ''

        # Case 3: Job number with optional suffix (e.g., 24.60, 24-60, 24 60, 24,01, 24, 01, 22.82-W, 24.09 - Joey Restaurant)
        job_match = re.match(r'^(\d{2})[.\-,\s,]{1,3}(\d{2,3})([\s\-–]*[A-Za-z].*)?$', val)
        if job_match:
            part1 = job_match.group(1)
            part2 = job_match.group(2)
            return '', f'{part1}.{part2}', '', ''

        # Case 4: Everything else → Remarks
        return '', '', '', val

    except Exception as e:
        return '', '', '', value

def process_pdf_folder(folder_path, output_csv="outputs/excel_files/pixtral_po_results.csv"):
    # Check if folder exists
    if not os.path.exists(folder_path):
        print(f"ERROR: Error: Folder {folder_path} does not exist!")
        return
    
    api_key = os.environ.get("MISTRAL_API_KEY")
    if not api_key:
        raise ValueError("MISTRAL_API_KEY not set in environment.")

    client = Mistral(api_key=api_key)
    results = []
    
    # Count PDF files
    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
    print(f"FOLDER: Found {len(pdf_files)} PDF files in {folder_path}")
    
    if len(pdf_files) == 0:
        print("WARNING: No PDF files found. Creating empty results.")
        results.append({
            "file_name": "no_files_found",
            "extracted_po_number": "",
            "clean_po_number": ""
        })

    for file in pdf_files:
        full_path = os.path.join(folder_path, file)
        print(f"FILE: Processing {file}...")

        img_path = convert_pdf_to_image(full_path)
        if not img_path:
            results.append({"file_name": file, "extracted_po_number": "ERROR: image conversion failed", "clean_po_number": ""})
            continue

        base64_img = encode_image(img_path)
        if not base64_img:
            results.append({"file_name": file, "extracted_po_number": "ERROR: base64 encode failed", "clean_po_number": ""})
            continue

        raw_po = extract_po_from_image(base64_img, client)
        clean_po = clean_po_value(raw_po if raw_po.lower() not in ["not found", "not visible"] else "")

        results.append({
            "file_name": file,
            "extracted_po_number": raw_po,
            "clean_po_number": clean_po
        })

        os.remove(img_path)

    # SUCCESS: Create DataFrame from extracted results
    df = pd.DataFrame(results)
    
    # Debug: Print DataFrame info
    print(f"DATA: DataFrame shape: {df.shape}")
    print(f"INFO: DataFrame columns: {list(df.columns)}")
    print(f"FILE: DataFrame head:\n{df.head()}")
    
    # Check if extracted_po_number column exists, if not create it
    if 'extracted_po_number' not in df.columns:
        print("WARNING: Warning: extracted_po_number column not found. Creating empty column.")
        df['extracted_po_number'] = ''
    
    # SUCCESS: Perform PO Classification using extracted_po_number
    try:
        df[['PO_Number', 'Job_Number', 'WO_Number', 'Remarks']] = df['extracted_po_number'].apply(
            lambda x: pd.Series(classify_po(x))
        )
    except Exception as e:
        print(f"ERROR: Error during PO classification: {e}")
        # Create empty columns if classification fails
        df['PO_Number'] = ''
        df['Job_Number'] = ''
        df['WO_Number'] = ''
        df['Remarks'] = ''

    # SUCCESS: Save the full output to CSV
    df.to_csv(output_csv, index=False)
    print(f"\nSUCCESS: PO numbers with classification saved to {output_csv}")

pdf_folder = "data/raw_pdfs"
output_csv = "outputs/excel_files/pixtral_po_results.csv"
process_pdf_folder(pdf_folder, output_csv)

# Explicitly treat PO_Number and Job_Number as string to preserve formatting like '24.60'
df_po = pd.read_csv(output_csv, dtype={"PO_Number": str, "Job_Number": str})

def clean_identifier(val):
    try:
        if pd.isna(val) or val == "":
            return ""
        val = str(val).strip()

        # Fix for numbers like 13511.0 → "13511"
        if re.fullmatch(r"\d+\.0", val):
            val = val.split(".")[0]

        return val
    except:
        return str(val)

df_po["PO_Number"] = df_po["PO_Number"].apply(clean_identifier)
df_po["Job_Number"] = df_po["Job_Number"].apply(clean_identifier)
df_po["WO_Number"] = df_po["WO_Number"].apply(clean_identifier)

import os

def extract_info(value, mode, directory_path):
    """
    Simulate extraction by checking for a matching filename or content in files.
    Parameters:
    - value: the PO or Job number to search for.
    - mode: either "po" or "job".
    - directory_path: where the files are located.
    
    Returns:
    - Tuple: (match_string or None, file_name or None)
    """
    try:
        # Check if directory exists
        if not os.path.exists(directory_path):
            print(f"WARNING: Directory not accessible: {directory_path}")
            return (None, None)
            
        for filename in os.listdir(directory_path):
            if filename.endswith(".pdf") or filename.endswith(".txt"):
                if value in filename:
                    return (f"SUCCESS: Found in {filename}", filename)
        return (None, None)
    except Exception as e:
        return (f"ERROR: Error: {str(e)}", None)

po_verified_by = []
job_verified_by = []

for _, row in df_po.iterrows():
    match_found = False

    if pd.notna(row["PO_Number"]) and row["PO_Number"] != "":
        result = extract_info(row["PO_Number"], "po", PROJECTS_PATH)
        if result[0]:
            po_verified_by.append(result[0])
            job_verified_by.append("")
            match_found = True

    if not match_found and pd.notna(row["Job_Number"]) and row["Job_Number"] != "":
        result = extract_info(row["Job_Number"], "job", PROJECTS_PATH)
        if result[0]:
            job_verified_by.append(result[0])
            po_verified_by.append("")
            match_found = True

    if not match_found:
        po_verified_by.append("ERROR: Not Found")
        job_verified_by.append("ERROR: Not Found")

df_po["po_verified_by"] = po_verified_by
df_po["job_verified_by"] = job_verified_by

# Convert entire DataFrame to string and replace NaN with empty string
df_po = df_po.astype(str).replace("nan", "")

final_columns = [
    "file_name", "extracted_po_number", "clean_po_number",
    "PO_Number", "Job_Number", "WO_Number", "Remarks",
    "po_verified_by", "job_verified_by"
]

# Check which columns exist and create missing ones
missing_columns = [col for col in final_columns if col not in df_po.columns]
if missing_columns:
    print(f"WARNING: Warning: Missing columns: {missing_columns}")
    for col in missing_columns:
        df_po[col] = ""

# Only use columns that exist
available_columns = [col for col in final_columns if col in df_po.columns]
df_po[available_columns].to_csv(output_csv, index=False)
print("SUCCESS: Final file saved:", output_csv)
print(f"INFO: Available columns: {available_columns}")
df_po[available_columns].head()

# Create PO approval flag with absolute path
current_dir = os.path.dirname(os.path.abspath(__file__))
po_approval_flag_file = os.path.join(current_dir, "..", "outputs", "excel_files", "po_approval_needed.flag")

# Ensure the directory exists
os.makedirs(os.path.dirname(po_approval_flag_file), exist_ok=True)

with open(po_approval_flag_file, 'w') as f:
    f.write("po_approval_needed")
print("SUCCESS: Created PO approval flag:", po_approval_flag_file)

def wait_for_po_approval():
    """Wait for user PO approval before continuing"""
    print("SUCCESS: PO matches created. Waiting for user approval...")
    print(f"SUCCESS: Approval flag path: {po_approval_flag_file}")
    
    # Create PO approval flag
    with open(po_approval_flag_file, 'w') as f:
        f.write("po_approval_needed")
    
    # Wait for approval with timeout and better error handling
    max_wait_time = 3600  # 1 hour timeout
    start_time = time.time()
    
    while True:
        try:
            current_time = time.time()
            if current_time - start_time > max_wait_time:
                print("ERROR: Timeout waiting for PO approval (1 hour). Continuing anyway...")
                return True
            
            if not os.path.exists(po_approval_flag_file):
                print("SUCCESS: PO approval received. Continuing with processing...")
                return True
            
            # Print progress every 30 seconds
            elapsed = int(current_time - start_time)
            if elapsed % 30 == 0:
                print(f"INFO: Still waiting for PO approval... ({elapsed} seconds elapsed)")
            
            time.sleep(1)  # Check every second
            
        except Exception as e:
            print(f"ERROR: Error checking PO approval status: {e}")
            time.sleep(1)

# Wait for PO approval
wait_for_po_approval()

# IMPORTANT: Re-read the PO CSV file to get the updated PO information
print("SUCCESS: Reading updated PO information from CSV file...")
updated_pixtral_df = pd.read_csv("outputs/excel_files/pixtral_po_results.csv")
print(f"SUCCESS: Loaded {len(updated_pixtral_df)} updated PO matches from CSV")

import pandas as pd
import json
import os

# === File paths ===
vendor_csv_path = "data/Vendor_List.csv"
pixtral_csv_path = "outputs/excel_files/pixtral_po_results.csv"
json_folder = "data/processed"

# === Load CSVs ===
vendor_df = pd.read_csv(vendor_csv_path)
vendor_df["Vendor_Code"] = vendor_df["Vendor_Code"].astype(str).str.strip()

# Use the updated PO data instead of the original
pixtral_df = updated_pixtral_df
pixtral_df["file_base"] = pixtral_df["file_name"].apply(lambda x: os.path.splitext(x)[0])

# === Process each JSON in the folder ===
for file in os.listdir(json_folder):
    if not file.endswith(".json"):
        continue

    json_path = os.path.join(json_folder, file)
    file_base = os.path.splitext(file)[0]

    # Load JSON
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    vendor_code = str(data.get("Vendor_Code", "")).strip()
    if not vendor_code:
        print(f"ERROR: Skipping {file} — Vendor_Code missing")
        continue

    # Look up vendor info
    vendor_row = vendor_df[vendor_df["Vendor_Code"] == vendor_code]
    if vendor_row.empty:
        print(f"ERROR: Skipping {file} — Vendor_Code not found in Vendor_List")
        continue

    vendor_info = vendor_row.iloc[0]
    gl_normal = str(vendor_info.get("Distribution_GL_Account", "")).strip()
    gl_wo = str(vendor_info.get("WO_GL_Codes", "")).strip()

    # Look up corresponding row in pixtral results
    pixtral_row = pixtral_df[pixtral_df["file_base"] == file_base]
    if pixtral_row.empty:
        print(f"ERROR: Skipping {file} — Not found in pixtral_po_results.csv")
        continue

    pixtral_row = pixtral_row.iloc[0]

    # Extract decision fields
    has_po_or_job = pd.notna(pixtral_row.get("PO_Number")) or pd.notna(pixtral_row.get("Job_Number"))
    has_wo_or_remark = pd.notna(pixtral_row.get("WO_Number")) or pd.notna(pixtral_row.get("Remarks"))

    # Extract and normalize remarks
    remarks = str(pixtral_row.get("Remarks", "")).strip().lower()

    # === If only one of the GL values is available, use it directly ===
    if gl_normal and not gl_wo:
        data["Distribution_GL_Account"] = str(int(float(gl_normal)))
    elif gl_wo and not gl_normal:
        data["Distribution_GL_Account"] = str(int(float(gl_wo)))
    
    # === If both GL values are present, apply conditional logic ===
    else:
        if remarks in {"shop", "stock", "shop stock", "shop fab", "shop sab"}:
            data["Distribution_GL_Account"] = "1200"  # Rule 1
        elif has_po_or_job:
            data["Distribution_GL_Account"] = str(int(float(gl_normal)))
            data["Phase_Code"] = vendor_info.get("Phase_Code", "")
            data["Cost_Type"] = vendor_info.get("Cost_Type", "")
        elif has_wo_or_remark:
            data["Distribution_GL_Account"] = str(int(float(gl_wo)))

    # Save updated JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print(f"SUCCESS: Enriched {file}")

# !pip install pymupdf xlrd

import os
import re
import pandas as pd
import fitz  # PyMuPDF
from tqdm import tqdm

pdf_folder = PROJECTS_PATH if PROJECTS_PATH else "data/network_fallback/"
input_path = "outputs/excel_files/pixtral_po_results.csv"
output_path = "outputs/excel_files/po_verified.csv"
routing_path = "data/Routing_Code.xlsx"
pm_file_path = "data/Job_Listing_by_PM.xls"

df_po = pd.read_csv(input_path)

def clean_number(val):
    try:
        return str(int(float(val))) if float(val).is_integer() else str(val)
    except:
        return str(val)

df_po["PO_Number"] = df_po["PO_Number"].apply(clean_number)
df_po["Job_Number"] = df_po["Job_Number"].apply(clean_number)
df_po["WO_Number"] = df_po["WO_Number"].apply(clean_number)

# Cache for PDF content to avoid re-processing the same files
pdf_cache = {}

def extract_info_optimized(identifiers, id_types, pdf_folder):
    """
    Optimized function that processes all identifiers at once using PDF caching
    """
    if not os.path.exists(pdf_folder):
        print(f"ERROR: Network path not accessible: {pdf_folder}")
        return {identifier: (None, "", "", "") for identifier in identifiers}
    
    # Get all PDF files once
    pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]
    print(f"INFO: Processing {len(pdf_files)} PDF files for {len(identifiers)} identifiers...")
    
    # Initialize results
    results = {identifier: (None, "", "", "") for identifier in identifiers}
    
    # Process each PDF file once and cache the results
    for pdf_file in tqdm(pdf_files, desc="Processing PDFs"):
        pdf_path = os.path.join(pdf_folder, pdf_file)
        
        try:
            # Check if PDF is already cached
            if pdf_file in pdf_cache:
                pdf_data = pdf_cache[pdf_file]
            else:
                # Process PDF and cache the result
                doc = fitz.open(pdf_path)
                pdf_text = ""
                for page in doc:
                    pdf_text += page.get_text()
                doc.close()
                
                # Cache the extracted text
                pdf_cache[pdf_file] = pdf_text
            
            pdf_text = pdf_cache[pdf_file]
            
            # Check all identifiers against this PDF
            for identifier, id_type in zip(identifiers, id_types):
                if results[identifier][0] is not None:
                    continue  # Already found a match for this identifier
                
                # Create pattern for this identifier
                pattern = re.compile(rf"{'Purchase Order' if id_type == 'po' else 'Job'}[:\s]*{re.escape(str(identifier))}", re.IGNORECASE)
                
                if pattern.search(pdf_text):
                    # Extract ordered_by and distribution_code
                    ordered_by_pattern = re.compile(r"Ordered By:\s*(.+)", re.IGNORECASE)
                    distribution_pattern = re.compile(r"\d{4}\s+([EMS])\b")
                    
                    ordered_by = ""
                    distribution_code = ""
                    
                    if (m := ordered_by_pattern.search(pdf_text)):
                        ordered_by = m.group(1).strip()
                    if (d := distribution_pattern.findall(pdf_text)):
                        distribution_code = d[0]
                    
                    results[identifier] = (pdf_file, ordered_by, distribution_code)
                    
        except Exception as e:
            print(f"ERROR: Error processing {pdf_file}: {e}")
            continue
    
    return results

# Get unique identifiers and their types
po_identifiers = []
job_identifiers = []
po_types = []
job_types = []

for _, row in df_po.iterrows():
    if pd.notna(row["PO_Number"]) and row["PO_Number"] != "":
        po_identifiers.append(row["PO_Number"])
        po_types.append("po")
    
    if pd.notna(row["Job_Number"]) and row["Job_Number"].strip() != "":
        job_identifiers.append(row["Job_Number"])
        job_types.append("job")

# Process all identifiers at once
print(f"INFO: Processing {len(po_identifiers)} PO numbers and {len(job_identifiers)} Job numbers...")

all_identifiers = po_identifiers + job_identifiers
all_types = po_types + job_types

if all_identifiers:
    results = extract_info_optimized(all_identifiers, all_types, pdf_folder)
else:
    results = {}

# Create lookup dictionaries for fast access
po_results = {po_id: results.get(po_id, (None, "", "", "")) for po_id in po_identifiers}
job_results = {job_id: results.get(job_id, (None, "", "", "")) for job_id in job_identifiers}

# Build the final arrays
po_verified_by = []
job_verified_by = []
ordered_by_final = []
distribution_code_final = []

for _, row in df_po.iterrows():
    po_verified = ""
    job_verified = ""
    ordered_by = ""
    dist_code = ""
    
    # Check PO Number
    if pd.notna(row["PO_Number"]) and row["PO_Number"] != "":
        if row["PO_Number"] in po_results:
            result = po_results[row["PO_Number"]]
            if result[0]:
                po_verified = result[0]
                ordered_by = result[1]
                dist_code = result[2]
    
    # Check Job Number (only if PO not found)
    if not po_verified and pd.notna(row["Job_Number"]) and row["Job_Number"].strip() != "":
        if row["Job_Number"] in job_results:
            result = job_results[row["Job_Number"]]
            if result[0]:
                job_verified = result[0]
                ordered_by = result[1]
                dist_code = result[2]
    
    # Set default values if nothing found
    if not po_verified and not job_verified:
        po_verified = "ERROR: Not Found"
        job_verified = "ERROR: Not Found"
    
    po_verified_by.append(po_verified)
    job_verified_by.append(job_verified)
    ordered_by_final.append(ordered_by)
    distribution_code_final.append(dist_code)

df_po["po_verified_by"] = po_verified_by
df_po["job_verified_by"] = job_verified_by
df_po["ordered_by"] = ordered_by_final
df_po["distribution_code"] = distribution_code_final

# Clear the cache to free memory
pdf_cache.clear()

print("INFO: PDF verification completed. Processing additional data...")

# Process job numbers from PO files (if needed)
job_number_pattern = re.compile(r"Job[:\s]*([\d\.]+)", re.IGNORECASE)

for i, row in tqdm(df_po.iterrows(), desc="Extracting job numbers from PO files", total=len(df_po)):
    job_num = row["Job_Number"].strip()
    po_file = row["po_verified_by"].strip()
    
    if (not job_num or job_num.lower() in ["", "nan", "none"]) and \
       (po_file and not po_file.startswith("ERROR:")):
        pdf_path = os.path.join(pdf_folder, po_file)
        try:
            # Use cached PDF if available
            if po_file in pdf_cache:
                pdf_text = pdf_cache[po_file]
            else:
                doc = fitz.open(pdf_path)
                pdf_text = ""
                for page in doc:
                    pdf_text += page.get_text()
                doc.close()
                # Cache for potential reuse
                pdf_cache[po_file] = pdf_text
            
            match = job_number_pattern.search(pdf_text)
            if match:
                df_po.at[i, "Job_Number"] = match.group(1).strip()
        except Exception as e:
            print(f"ERROR: Error reading PDF {po_file}: {e}")

# Clear cache again
pdf_cache.clear()

# -------------------------------
# PM NAME LOOKUP AND REPLACE ORDERED_BY
# -------------------------------
df_pm = pd.read_excel(pm_file_path)

df_pm["Job #"] = df_pm["Job #"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
df_pm["Project Mgr."] = df_pm["Project Mgr."].astype(str).str.strip().str.upper()
df_po["Job_Number"] = df_po["Job_Number"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)

# Apply consistent xx.xx formatting
def format_job_number(val):
    try:
        num = float(val)
        return f"{num:.2f}"
    except:
        return val

df_po["Job_Number"] = df_po["Job_Number"].apply(format_job_number)

code_to_name = {
    "RAORAK": "Rakesh Rao",
    "TANARV": "Arvind Tandel",
    "TANMUK": "Mukesh Tandel",
    "TANNIR": "Niraj Tandel",
    "TANJIG": "Jignesh Tandel",
    "DANDRO": "Dan Droubay",
    "PATSHI": "Shirish Patel",
    "ANDSTE": "Steven Anderson",
    "ANIABH": "Anil Abhay"
}

# Build lookup for PM Name
pm_dict = {
    row["Job #"]: row["Project Mgr."]
    for _, row in df_pm.iterrows()
}

print("INFO: Processing PM assignments...")

# Replace ordered_by with PM Full Name
df_po["ordered_by"] = df_po["Job_Number"].map(pm_dict).map(code_to_name).fillna("")

# -------------------------------
# ROUTING CODE BASED ON NEW ordered_by
# -------------------------------
print("INFO: Processing routing codes...")
df_routing = pd.read_excel(routing_path)

df_routing["Ordered By"] = df_routing["Ordered By"].astype(str).str.strip().str.upper()
df_routing["Distribution"] = df_routing["Distribution"].astype(str).str.strip().str.upper()
df_routing["Code"] = df_routing["Code"].astype(str).str.strip()

# Build routing lookup dictionary
routing_dict = {}
for _, row in df_routing.iterrows():
    ordered_by = row["Ordered By"]
    distribution = row["Distribution"]
    code = row["Code"]
    
    if ordered_by and distribution and code:
        key = (ordered_by, distribution)
        routing_dict[key] = code

print(f"INFO: Built routing lookup with {len(routing_dict)} entries")

# Apply routing codes
routing_codes = []
for _, row in tqdm(df_po.iterrows(), desc="Applying routing codes", total=len(df_po)):
    ordered_by = row["ordered_by"].strip().upper()
    distribution_code = row["distribution_code"].strip().upper()
    
    if ordered_by and distribution_code:
        key = (ordered_by, distribution_code)
        routing_code = routing_dict.get(key, "")
    else:
        routing_code = ""
    
    routing_codes.append(routing_code)

df_po["routing_code"] = routing_codes

# -------------------------------
# SAVE THE FINAL PO VERIFIED FILE
# -------------------------------
print("INFO: Saving final po_verified.csv file...")

# Ensure all required columns exist
required_columns = [
    "file_name", "extracted_po_number", "clean_po_number",
    "PO_Number", "Job_Number", "WO_Number", "Remarks",
    "po_verified_by", "job_verified_by", "ordered_by", 
    "distribution_code", "routing_code"
]

for col in required_columns:
    if col not in df_po.columns:
        df_po[col] = ""

# Save the final file
df_po[required_columns].to_csv(output_path, index=False)
print(f"SUCCESS: Final po_verified.csv saved to {output_path}")
print(f"INFO: File contains {len(df_po)} rows with {len(required_columns)} columns")

# Show summary statistics
print("\n" + "="*50)
print("PROCESSING SUMMARY")
print("="*50)
print(f"Total records processed: {len(df_po)}")
print(f"PO numbers found: {len([x for x in df_po['po_verified_by'] if x and not x.startswith('ERROR')])}")
print(f"Job numbers found: {len([x for x in df_po['job_verified_by'] if x and not x.startswith('ERROR')])}")
print(f"Routing codes assigned: {len([x for x in df_po['routing_code'] if x])}")
print("="*50)

# Cell 1: Import Libraries
# !pip install pymupdf opencv-python pytesseract numpy pandas

# Standard Imports
import os
import fitz 
import cv2
import pandas as pd
import pytesseract
import numpy as np
import shutil  # For cleanup

# Set your Tesseract path
current_dir = os.path.dirname(os.path.abspath(__file__))
# Option 1: Use Windows Tesseract installation (recommended)
# TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# Option 2: Use local models folder (if you install there)
# TESSERACT_PATH = os.path.join(current_dir, "..", "models", "tesseract-5.3.3", "tesseract.exe")
# Option 3: Custom installation path (if you installed elsewhere)
# TESSERACT_PATH = r"C:\Users\raj\Desktop\DO NOT TOUCH\invoice_extraction_project\models\Tesseract-OCR\tesseract.exe"
# Option 4: Current Tesseract installation
TESSERACT_PATH = os.path.join(current_dir, "..", "models", "tesseract", "tesseract.exe")
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

import re

def clean_base_name(filename):
    """
    Removes _page_# or _part_# suffix and file extension to standardize base names.
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    base = re.sub(r'(_page_\d+|_part_\d+)$', '', base)
    return base

# Cell 2: Load Logo Template
logo_template_path = r"data/Picture1.jpg"  # <-- Change to your actual logo path
logo_template = cv2.imread(logo_template_path, 0)

# Cell 3: Detect Logo Using Template Matching
def page_has_logo_template(image_path, template=logo_template, threshold=0.20):
    image = cv2.imread(image_path, 0)
    if image is None or template is None:
        return False

    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, _ = cv2.minMaxLoc(result)

    return max_val >= threshold

def pdf_to_images_logo_filtered(pdf_folder, output_folder, valid_files_list):
    os.makedirs(output_folder, exist_ok=True)
    valid_image_paths = []

    for file in os.listdir(pdf_folder):
        if file not in valid_files_list:
            continue

        if file.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, file)
            doc = fitz.open(pdf_path)
            base_name = os.path.splitext(file)[0]

            for i, page in enumerate(doc):
                img_path = os.path.join(output_folder, f"{base_name}_page_{i+1}.png")
                pix = page.get_pixmap(dpi=250)
                pix.save(img_path)

                if page_has_logo_template(img_path):
                    valid_image_paths.append(img_path)
                else:
                    os.remove(img_path)

    return valid_image_paths

# Cell 5: Header and Total Detection with OCR
HEADER_KEYWORDS = ["Seq", "Part", "Description", "U/M", "Quantity", "Unit", "Disc%", "Tax%", "Extension", "Distribution"]

def find_header_y(image_path):
    image = cv2.imread(image_path)
    data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    header_y = None

    for i, text in enumerate(data['text']):
        if text.strip() in HEADER_KEYWORDS:
            y = data['top'][i]
            if header_y is None or y < header_y:
                header_y = y
    return header_y

def find_total_y(image_path):
    image = cv2.imread(image_path)
    data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
    total_y = None

    for i, text in enumerate(data['text']):
        if "Total" in text:
            total_y = data['top'][i]
            break  # Stop after finding first Total
    return total_y

def extract_text_until_total(image_path):
    image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    text = pytesseract.image_to_string(gray, config='--psm 6')
    lines = text.splitlines()

    for line in lines:
        if "Total" in line:
            print(f"Detected Total Line: {line.strip()}")
            break

def process_po_folder(pdf_folder, image_folder, cropped_folder, text_output_folder):
    os.makedirs(image_folder, exist_ok=True)
    os.makedirs(cropped_folder, exist_ok=True)
    os.makedirs(text_output_folder, exist_ok=True)

    df_po = pd.read_csv("outputs/excel_files/po_verified.csv")
    valid_files = df_po["po_verified_by"].dropna()
    valid_files = valid_files[~valid_files.str.contains("ERROR: Not Found", na=False)]
    valid_files_list = valid_files.unique().tolist()

    valid_image_paths = pdf_to_images_logo_filtered(pdf_folder, image_folder, valid_files_list)
    grouped_images = {}

    for img_path in valid_image_paths:
        pdf_base_name = clean_base_name(img_path)
        grouped_images.setdefault(pdf_base_name, []).append(img_path)

    for pdf_base_name, image_list in grouped_images.items():
        image_list.sort()
        collecting, header_found = False, False
        table_parts = []

        for img_path in image_list:
            image = cv2.imread(img_path)
            header_y = find_header_y(img_path)
            total_y = find_total_y(img_path)

            if header_y is not None and not header_found:
                cropped = image[header_y:, :]
                table_parts.append(cropped)
                header_found = True
                collecting = True
                continue

            if collecting and total_y is not None:
                cropped = image[:total_y, :]
                table_parts.append(cropped)
                break

            elif collecting:
                table_parts.append(image)

        if table_parts:
            max_width = max(part.shape[1] for part in table_parts)
            resized_parts = []

            for part in table_parts:
                h, w = part.shape[:2]
                if w < max_width:
                    scale = max_width / w
                    resized = cv2.resize(part, (max_width, int(h * scale)), interpolation=cv2.INTER_LINEAR)
                    resized_parts.append(resized)
                else:
                    resized_parts.append(part)

            combined = np.vstack(resized_parts)
            cropped_path = os.path.join(cropped_folder, f"{pdf_base_name}.png")
            cv2.imwrite(cropped_path, combined)

            gray = cv2.cvtColor(combined, cv2.COLOR_BGR2GRAY)
            text = pytesseract.image_to_string(gray, config='--psm 6')

            filtered_output = ""
            for line in text.splitlines():
                if "!" in line:
                    filtered_output += line.replace(",", "") + "\n"

            text_file = os.path.join(text_output_folder, f"{pdf_base_name}.txt")
            with open(text_file, "w", encoding="utf-8") as f:
                f.write(filtered_output)

            print(f"Processed PO: {pdf_base_name}")

# Cell 7: Example Run (Updated paths)
process_po_folder(
    pdf_folder=PROJECTS_PATH if PROJECTS_PATH else "data/network_fallback/",
image_folder=r"data/image_of_pos",
cropped_folder=r"data/cropped_images",
text_output_folder=r"data/po_ocr_output"
)

import os
import re
import pandas as pd

input_folder = "data/po_ocr_output"
output_path = "data/po_ocr_extracted/final_extracted.csv"

df_po = pd.read_csv("outputs/excel_files/po_verified.csv")
valid_files = df_po["po_verified_by"].dropna()
valid_files = valid_files[~valid_files.str.contains("ERROR: Not Found", na=False)]
valid_basenames = [clean_base_name(f) for f in valid_files.unique()]

results = []

for file in os.listdir(input_folder):
    if not file.lower().endswith(".txt"):
        continue

    txt_basename = clean_base_name(file)

    if txt_basename not in valid_basenames:
        continue  # Skip unrelated txt files

    with open(os.path.join(input_folder, file), "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = re.sub(r'\s+', ' ', line.strip())

        seq = re.match(r"^(\d{3})", line).group(1) if re.match(r"^(\d{3})", line) else ""
        distribution_match = re.search(r"(\d{4,5}\s?[A-Z]{1,3})$", line)
        distribution = ""
        dist_match = re.search(r"\b([1-9]\d{3})\b$", line)
        if dist_match:
            distribution = dist_match.group(1)

        part_match = re.search(r"(!.*?)(?=\d+\.\d{2,4})", line)
        part_number = part_match.group(1).strip() if part_match else ""

        start = part_match.end() if part_match else 0
        end = distribution_match.start() if distribution_match else len(line)
        remaining = line[start:end].strip()

        quantity = ""
        unit_cost = ""
        tax_percent = ""
        extension = ""

        number_matches = list(re.finditer(r"\d+(?:,\d{3})*(?:\.\d{2,4})", remaining))

        if number_matches:
            extension_val = number_matches[-1].group(0)
            extension = extension_val

            if len(number_matches) >= 2:
                before_ext = number_matches[-2].group(0)
                if float(before_ext.replace(",", "")) == float(extension.replace(",", "")):
                    unit_cost = before_ext
                else:
                    tax_percent = before_ext
                    if len(number_matches) >= 3:
                        potential_uc = number_matches[-3].group(0)
                        if float(potential_uc.replace(",", "")) < float(extension.replace(",", "")):
                            unit_cost = potential_uc

            if unit_cost:
                unit_cost_match = re.search(re.escape(unit_cost), remaining)
                if unit_cost_match:
                    quantity_section = remaining[:unit_cost_match.start()].strip()
                    quantity_tokens = quantity_section.split()
                    if quantity_tokens:
                        quantity = quantity_tokens[-1]
            else:
                qty_end = number_matches[-3].start() if tax_percent and len(number_matches) >= 3 else (
                    number_matches[-2].start() if len(number_matches) >= 2 else -1
                )
                if qty_end > 0:
                    quantity_section = remaining[:qty_end].strip()
                    quantity_tokens = quantity_section.split()
                    if quantity_tokens:
                        quantity = quantity_tokens[-1]

        results.append({
            "Seq": seq,
            "Part Number": part_number,
            "Quantity": quantity,
            "Unit Cost": unit_cost,
            "Tax %": tax_percent,
            "Extension": extension,
            "Distribution": distribution,
            "Source File": file
        })

# Save output
os.makedirs(os.path.dirname(output_path), exist_ok=True)
df = pd.DataFrame(results)
df.to_csv(output_path, index=False)
print(f"SUCCESS: Extraction completed. Results saved to {output_path}")

import os
import json
import pandas as pd
import re
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

json_folder = 'data/processed'  # Change path if needed
all_data = []

for file_name in tqdm(os.listdir(json_folder)):
    if file_name.endswith('.json'):
        with open(os.path.join(json_folder, file_name), 'r') as f:
            data = json.load(f)
            if isinstance(data, dict):
                data['file_name'] = file_name.replace('.json', '.pdf')
                all_data.append(data)

df = pd.DataFrame(all_data)

# Ensure columns exist even if missing from some JSONs
for col in ['Amount_Before_Taxes']:
    if col not in df.columns:
        df[col] = ''

po_df = pd.read_csv('outputs/excel_files/pixtral_po_results.csv')
po_df['file_name'] = po_df['file_name'].str.lower().str.strip()
df['file_name'] = df['file_name'].str.lower().str.strip()

# SUCCESS: Merge in all needed columns
df = df.merge(
    po_df[['file_name', 'extracted_po_number', 'PO_Number', 'Job_Number','WO_Number', 'Remarks']],
    on='file_name',
    how='left'
)

# Define cleaning functions before they are used
def clean_invoice_amount(value):
    try:
        # Remove currency symbols and commas, but preserve decimal points
        cleaned = str(value).replace('$', '').replace(',', '').strip()
        # Ensure proper decimal formatting
        if cleaned and '.' in cleaned:
            # Convert to float and back to string to ensure proper decimal format
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        elif cleaned:
            # If no decimal point, assume it's a whole number
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        return cleaned
    except:
        return value

def clean_amount_before_taxes(value):
    try:
        # Remove currency symbols and commas, but preserve decimal points
        cleaned = str(value).replace('$', '').replace(',', '').strip()
        # Ensure proper decimal formatting
        if cleaned and '.' in cleaned:
            # Convert to float and back to string to ensure proper decimal format
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        elif cleaned:
            # If no decimal point, assume it's a whole number
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        return cleaned
    except:
        return value

def clean_tax_amount(value):
    try:
        # Remove currency symbols and commas, but preserve decimal points
        cleaned = str(value).replace('$', '').replace(',', '').strip()
        # Ensure proper decimal formatting
        if cleaned and '.' in cleaned:
            # Convert to float and back to string to ensure proper decimal format
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        elif cleaned:
            # If no decimal point, assume it's a whole number
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        return cleaned
    except:
        return value

def clean_shipping_charges(value):
    try:
        # Remove currency symbols and commas, but preserve decimal points
        cleaned = str(value).replace('$', '').replace(',', '').strip()
        # Ensure proper decimal formatting
        if cleaned and '.' in cleaned:
            # Convert to float and back to string to ensure proper decimal format
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        elif cleaned:
            # If no decimal point, assume it's a whole number
            float_val = float(cleaned)
            return f"{float_val:.2f}"
        return cleaned
    except:
        return value

def determine_invoice_type(amount_str):
    try:
        amount = float(amount_str)
        return 'I' if amount >= 0 else 'C'
    except:
        return ''

def clean_amount_for_credit(amount_str, invoice_type):
    """Clean amount and make it positive if it's a credit invoice"""
    try:
        # Remove currency symbols and commas, but preserve decimal points
        cleaned = str(amount_str).replace('$', '').replace(',', '').strip()
        if cleaned:
            float_val = float(cleaned)
            if invoice_type == 'C':
                # For credit invoices, make amount positive (remove minus sign)
                return f"{abs(float_val):.2f}"
            else:
                return f"{float_val:.2f}"
        return cleaned
    except:
        return amount_str



# Clean amount fields to preserve decimal points
df['Invoice_Amount'] = df['Invoice_Amount'].apply(clean_invoice_amount)
df['Amount_Before_Taxes'] = df['Amount_Before_Taxes'].apply(clean_amount_before_taxes)
df['Tax_Amount'] = df['Tax_Amount'].apply(clean_tax_amount)

# Handle Shipping_Charges from JSON files
if 'Shipping_Charges' in df.columns:
    df['Shipping_Charges'] = df['Shipping_Charges'].apply(clean_shipping_charges)
else:
    df['Shipping_Charges'] = ''

df['Invoice_Type'] = df['Invoice_Amount'].apply(determine_invoice_type)

# Now clean amounts again, making them positive for credit invoices
df['Invoice_Amount'] = df.apply(lambda row: clean_amount_for_credit(row['Invoice_Amount'], row['Invoice_Type']), axis=1)
df['Amount_Before_Taxes'] = df.apply(lambda row: clean_amount_for_credit(row['Amount_Before_Taxes'], row['Invoice_Type']), axis=1)
df['Tax_Amount'] = df.apply(lambda row: clean_amount_for_credit(row['Tax_Amount'], row['Invoice_Type']), axis=1)

# Function to handle multiple date formats and convert to MM/DD/YY
def convert_date_to_mm_dd_yy(date_str):
    if pd.isna(date_str) or date_str == "" or date_str == "NaT":
        return ""
    
    try:
        # Try different date formats
        date_formats = [
            "%m/%d/%Y",  # 07/23/2025
            "%m/%d/%y",  # 07/23/25
            "%Y-%m-%d",  # 2025-07-23
            "%m-%d-%Y",  # 07-23-2025
            "%m-%d-%y",  # 07-23-25
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = pd.to_datetime(date_str, format=fmt)
                return parsed_date.strftime("%m/%d/%y")
            except:
                continue
        
        # If specific formats fail, try pandas automatic parsing
        parsed_date = pd.to_datetime(date_str, errors="coerce")
        if not pd.isna(parsed_date):
            return parsed_date.strftime("%m/%d/%y")
        
        return ""
    except:
        return ""

# Convert Invoice_Date and GL_Date with flexible format handling
df["Invoice_Date"] = df["Invoice_Date"].apply(convert_date_to_mm_dd_yy)
df["GL_Date"] = df["GL_Date"].apply(convert_date_to_mm_dd_yy)

# Ensure both Invoice_Date and GL_Date are filled
for index, row in df.iterrows():
    invoice_date = row["Invoice_Date"]
    gl_date = row["GL_Date"]
    
    # If Invoice_Date is not empty, ensure GL_Date matches it
    if pd.notna(invoice_date) and invoice_date != "" and invoice_date != "NaT":
        # Always set GL_Date to match Invoice_Date (they should be the same)
        df.at[index, "GL_Date"] = invoice_date
    # If Invoice_Date is empty but GL_Date has value, copy GL_Date to Invoice_Date
    elif (pd.isna(invoice_date) or invoice_date == "" or invoice_date == "NaT") and (pd.notna(gl_date) and gl_date != "" and gl_date != "NaT"):
        df.at[index, "Invoice_Date"] = gl_date
    # If both are empty, try to get from JSON files as fallback
    else:
        # This will be handled by the JSON update logic if needed
        pass

print(f"SUCCESS: Ensured Invoice_Date and GL_Date are filled for {len(df)} records")

def compute_shipping(row):
    # First try to use shipping charges from JSON extraction
    if 'Shipping_Charges' in row and pd.notna(row['Shipping_Charges']) and str(row['Shipping_Charges']).strip() != '':
        try:
            shipping_charge = str(row['Shipping_Charges']).strip()
            # For credit invoices, make shipping charge positive
            if row['Invoice_Type'] == 'C':
                try:
                    return str(abs(float(shipping_charge)))
                except:
                    return shipping_charge
            return shipping_charge
        except:
            pass
    
    # Fallback: calculate shipping as difference
    try:
        invoice_amt = float(row['Invoice_Amount'])
        tax_amt = float(row['Tax_Amount'])
        before_tax_amt = float(row['Amount_Before_Taxes'])
        diff = round(invoice_amt - (tax_amt + before_tax_amt), 2)
        # For credit invoices, make shipping charge positive
        if row['Invoice_Type'] == 'C':
            return '' if abs(diff) < 0.01 else str(abs(diff))
        return '' if abs(diff) < 0.01 else str(diff)
    except:
        return ''

df['Shipping_Charges'] = df.apply(compute_shipping, axis=1)

# Set batch code to today's date (when processing starts)
from datetime import datetime
today_date = datetime.now().strftime('%Y-%m-%d')
df['Batch_Code'] = today_date

df['Company_Code'] = 'AA1'

# Ensure Job_Number column exists even if empty
if 'Job_Number' not in df.columns:
    df['Job_Number'] = ''

if 'Amount_Before_Taxes' not in df.columns:
    df['Amount_Before_Taxes'] = ''

# Optional drop: extracted_po_number
df.drop(columns=['extracted_po_number'], inplace=True)

# Drop 'source_file' if exists
df.drop(columns=['source_file'], errors='ignore', inplace=True)

# Reorder columns for output
columns_order = [
    'Company_Code', 'GL_Date', 'Batch_Code',
    'Vendor_Code', 'Vendor_Name', 'Invoice_Type',
    'Invoice_Number', 'Invoice_Date', 'Invoice_Amount', 'Tax_Amount',
    'Shipping_Charges', 'Amount_Before_Taxes', 'PO_Number', 'Job_Number',
    'WO_Number', 'Distribution_GL_Account', 'Phase_Code', 'Cost_Type',
    'Remarks', 'file_name'
]

df = df[[col for col in columns_order if col in df.columns]]

output_path = 'outputs/excel_files/final_invoice_data.xlsx'
os.makedirs(os.path.dirname(output_path), exist_ok=True)

# Save Excel first
df.to_excel(output_path, index=False)

# Adjust column widths
wb = load_workbook(output_path)
ws = wb.active

for i, col in enumerate(ws.iter_cols(), 1):  # 1-based index
    max_length = 0
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # limit to 50 characters wide
    ws.column_dimensions[get_column_letter(i)].width = adjusted_width

wb.save(output_path)
print("SUCCESS: Final Excel saved with PO classification and auto-sized columns:", output_path)

import pandas as pd
import numpy as np
import os

# Load required files
invoice_df = pd.read_excel("outputs/excel_files/final_invoice_data.xlsx")
po_verified_df = pd.read_csv("outputs/excel_files/po_verified.csv")
tax_code_df = pd.read_excel("outputs/excel_files/filtered_tax_percentages_with_codes.xlsx")

# Check if PO lines file exists and has data
po_lines_file = "data/po_ocr_extracted/final_extracted.csv"
if os.path.exists(po_lines_file):
    try:
        po_lines_df = pd.read_csv(po_lines_file)
        if po_lines_df.empty or len(po_lines_df.columns) == 0:
            print("WARNING: No invoice with PO is available - PO lines file is empty")
            po_lines_df = pd.DataFrame()  # Create empty DataFrame
        else:
            print(f"SUCCESS: Loaded PO lines data with {len(po_lines_df)} rows")
    except Exception as e:
        print(f"WARNING: Error reading PO lines file: {e}")
        print("WARNING: No invoice with PO is available")
        po_lines_df = pd.DataFrame()  # Create empty DataFrame
else:
    print("WARNING: No invoice with PO is available - PO lines file does not exist")
    po_lines_df = pd.DataFrame()  # Create empty DataFrame

# Clean filenames to base names (no extensions)
po_verified_df["Base_File_Name"] = po_verified_df["po_verified_by"].apply(lambda x: os.path.splitext(str(x))[0])

# Only process PO lines if DataFrame is not empty
if not po_lines_df.empty and "Source File" in po_lines_df.columns:
    po_lines_df["Base_File_Name"] = po_lines_df["Source File"].apply(lambda x: os.path.splitext(str(x))[0])
else:
    print("WARNING: Skipping PO lines processing - no valid PO data available")

# Output container
output_rows = []

# Process each invoice
for _, inv_row in invoice_df.iterrows():
    
    po_number = int(inv_row["PO_Number"]) if not pd.isna(inv_row["PO_Number"]) else None
    if not po_number:
        continue

    invoice_number = inv_row["Invoice_Number"]
    
    # Format dates as MM/DD/YY
    invoice_date = pd.to_datetime(inv_row["Invoice_Date"]).strftime("%m/%d/%y") if not pd.isna(inv_row["Invoice_Date"]) else ""
    gl_date = pd.to_datetime(inv_row["GL_Date"]).strftime("%m/%d/%y") if not pd.isna(inv_row["GL_Date"]) else ""

    total_amount = float(str(inv_row["Invoice_Amount"]).replace(",", ""))
    tax_amount = float(str(inv_row["Tax_Amount"]).replace(",", "")) if not pd.isna(inv_row["Tax_Amount"]) else 0.00
    amount_before_tax = float(str(inv_row["Amount_Before_Taxes"]).replace(",", "")) if not pd.isna(inv_row["Amount_Before_Taxes"]) else 0.00
    shipping_charges = float(str(inv_row["Shipping_Charges"]).replace(",", "")) if "Shipping_Charges" in inv_row and not pd.isna(inv_row["Shipping_Charges"]) else 0.00
    
    # Find source file
    po_match = po_verified_df.loc[po_verified_df["PO_Number"] == po_number]
    if po_match.empty:
        print(f"WARNING: No source file found for PO Number: {po_number}")
        continue
    
    source_file_base = po_match.iloc[0]["Base_File_Name"]
    routing_code = po_match.iloc[0]["routing_code"]
    
    # Find PO lines matching the source file
    if po_lines_df.empty or "Base_File_Name" not in po_lines_df.columns:
        print(f"WARNING: No PO lines data available for PO Number: {po_number}")
        continue
        
    po_lines = po_lines_df[po_lines_df["Base_File_Name"] == source_file_base].dropna(subset=["Unit Cost", "Extension"])
    
    if po_lines.empty:
        print(f"WARNING: No PO lines found for Source File: {source_file_base}")
        continue

    # Check if required columns exist
    required_columns = ["Extension", "Seq", "Part Number", "Quantity"]
    missing_columns = [col for col in required_columns if col not in po_lines.columns]
    if missing_columns:
        print(f"WARNING: Missing required columns in PO lines: {missing_columns}")
        continue
        
    # Best match by closest Extension (total amount)
    po_lines["Extension_Clean"] = po_lines["Extension"].astype(str).str.replace(",", "").astype(float)
    po_lines["Extension_Diff"] = po_lines["Extension_Clean"] - total_amount
    
    eligible_lines = po_lines[po_lines["Extension_Diff"] >= 0]
    
    if not eligible_lines.empty:
        best_line = eligible_lines.loc[eligible_lines["Extension_Diff"].idxmin()]
    else:
        po_lines["Extension_Diff_Abs"] = po_lines["Extension_Diff"].abs()
        best_line = po_lines.loc[po_lines["Extension_Diff_Abs"].idxmin()]

    seq = str(best_line["Seq"]).zfill(3)
    part_number = best_line["Part Number"]
    quantity = float(str(best_line["Quantity"]).replace(",", "")) if not pd.isna(best_line["Quantity"]) else 0.00
    
    # Tax % Calculation and Lookup
    if amount_before_tax and amount_before_tax != 0:
        tax_percent = round(((total_amount - amount_before_tax - shipping_charges) / amount_before_tax) * 100, 4)
    else:
        tax_percent = 0.00
    
    tax_code_lookup = tax_code_df.loc[round(tax_code_df["Tax %"], 4) == round(tax_percent, 4), "Tax Code"]
    tax_code = tax_code_lookup.iloc[0] if not tax_code_lookup.empty else f"{tax_percent:.4f}%"

    # Header Line
    header = [
        "H", po_number, invoice_date, invoice_number, invoice_date, gl_date,
        f"{total_amount:.2f}", "", "", "", "", "", "", routing_code, "", f"{tax_amount:.2f}"
    ]
    output_rows.append(header)

    # Detail Line (Product)
    detail = [
        "D", po_number, invoice_number, seq, part_number,
        f"{quantity:.2f}", f"{amount_before_tax:.2f}", tax_code, ""
    ]
    output_rows.append(detail)

    # Shipping Line (if present)
    if shipping_charges > 0:
        last_line = po_lines.tail(1).iloc[0]

        shipping_seq = str(last_line["Seq"]).zfill(3)
        shipping_part = last_line["Part Number"]
        shipping_quantity = float(str(last_line["Quantity"]).replace(",", "")) if not pd.isna(last_line["Quantity"]) else 0.00

        shipping_detail = [
            "D", po_number, invoice_number, shipping_seq,
            shipping_part, f"{shipping_quantity:.2f}", f"{shipping_charges:.2f}", "NT", ""
        ]
        output_rows.append(shipping_detail)

# Convert to DataFrame and save CSV (same as before)
output_df = pd.DataFrame(output_rows)
output_df.to_csv("outputs/excel_files/invoice_spectrum_format.csv", index=False, header=False)

# Save cleaned TXT with correct comma formatting (no trailing commas, but internal blanks preserved)
# Save TXT with exact field positions, including trailing commas
txt_path = "outputs/excel_files/invoice_spectrum_format.txt"
with open(txt_path, "w") as f:
    for row in output_rows:
        # Convert all fields to strings
        row_str = [str(x) for x in row]
        # Join and write exactly as-is (including empty strings)
        line = ",".join(row_str)
        f.write(line + "\n")

print("SUCCESS: Spectrum import files generated successfully (CSV and cleaned TXT).")

import os
import re
import math
import pandas as pd
from datetime import datetime

# === File paths ===
source_path = "outputs/excel_files/final_invoice_data.xlsx"
gl_item_code_path = "data/GL_Item_Code.csv"
out_txt_path = "outputs/excel_files/APInvoicesImport1.txt"

# === Helpers ===
def fmt_date_mmddyyyy(val):
    if pd.isna(val) or str(val).strip().lower() == "nan" or str(val).strip() == "":
        return ""
    if isinstance(val, (pd.Timestamp, datetime)):
        dt = val
    else:
        s = str(val).strip()
        try:
            num = float(s)
            dt = pd.to_datetime(num, unit="D", origin="1899-12-30", errors="coerce")
            if pd.isna(dt):
                raise ValueError
        except Exception:
            dt = None
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%y"):
                try:
                    dt = datetime.strptime(s, fmt)
                    break
                except Exception:
                    pass
            if dt is None:
                return ""
    return f"{dt.month:02d}{dt.day:02d}{dt.year:04d}"

def _sanitize_text(s: str) -> str:
    """Remove CSV-breaking chars and normalize whitespace."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace(",", " ").replace("\r", " ").replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def clean_str(v, maxlen=None):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        s = ""
    else:
        s = str(v).strip()
        if s.lower() == "nan":
            s = ""
    s = _sanitize_text(s)
    if maxlen is not None:
        return s[:maxlen]
    return s

def clean_num(n, decimals=2, allow_negative=True):
    if n is None or (isinstance(n, float) and math.isnan(n)) or str(n).strip().lower() == "nan":
        return ""
    try:
        x = float(str(n).replace(",", ""))
        if not allow_negative and x < 0:
            x = abs(x)
        return f"{x:.{decimals}f}"
    except Exception:
        return ""

def strip_dotzero(s):
    s = clean_str(s)
    return s[:-2] if s.endswith(".0") else s

def pad_or_trim_to(fields, expected_len):
    fields = list(fields)
    if len(fields) < expected_len:
        fields.extend([""] * (expected_len - len(fields)))
    elif len(fields) > expected_len:
        fields = fields[:expected_len]
    return fields

def normalize_phase_code(val: str) -> str:
    """
    Phase codes must be 4 digits, typically starting with 0 (e.g., 0320).
    Exception: '1000' remains '1000'.
    Input examples like '320', '0320', '320.0' all -> '0320'.
    """
    s = clean_str(val)
    if not s:
        return ""
    s = s.split(".")[0]              # drop decimal part if present
    s = re.sub(r"\D", "", s)         # keep digits only
    if not s:
        return ""
    if s == "1000":
        return "1000"
    s = s[-4:]                       # in case it's longer
    return s.zfill(4)

# === Load data ===
source_df = pd.read_excel(source_path)

# Build GL→Item map (robust to single-column CSV)
gl_item_df = pd.read_csv(gl_item_code_path, sep=",", header=0)
gl_item_df.columns = gl_item_df.columns.str.strip()
if gl_item_df.shape[1] == 1:
    gl_item_df = gl_item_df.iloc[:, 0].str.split(",", expand=True)
    gl_item_df.columns = ["G/L Code", "Item_Code"]
gl_item_df["G/L Code"] = gl_item_df["G/L Code"].astype(str).str.strip()
gl_item_df["Item_Code"] = gl_item_df["Item_Code"].astype(str).str.strip()
gl_to_item_map = dict(zip(gl_item_df["G/L Code"], gl_item_df["Item_Code"]))

# Defaults
for cc_col, default in [("Liability_Cost_Center","1000"), ("Expense_Cost_Center","1000")]:
    if cc_col not in source_df.columns:
        source_df[cc_col] = default

# === Skip rows with a PO_Number ===
def has_po(row):
    po = clean_str(row.get("PO_Number", ""))
    return po != ""

work_df = source_df[~source_df.apply(has_po, axis=1)].copy()

# === Precompute WO_Number and Item_Code per your rules ===
def compute_wo(row):
    remarks = clean_str(row.get("Remarks", "")).lower()
    gl_code = strip_dotzero(row.get("Distribution_GL_Account", ""))
    wo_val = strip_dotzero(row.get("WO_Number", ""))

    is_remarks_present = remarks not in ("", "nan")
    is_wo_present = wo_val not in ("", "nan")

    if is_wo_present:
        return "2025"
    if is_remarks_present:
        return "" if gl_code == "1200" else "2025"
    return ""

def compute_item_code(row):
    gl_code = strip_dotzero(row.get("Distribution_GL_Account", ""))
    ic = gl_to_item_map.get(gl_code, "")
    if ic:
        return ic if ic.startswith("!") else f"!{ic}"
    return ""

work_df["__WO_OUT__"] = work_df.apply(compute_wo, axis=1)
work_df["__ITEM_CODE__"] = work_df.apply(compute_item_code, axis=1)

# === Routing (header-level) ===
def compute_routing_for_group(df_group: pd.DataFrame) -> str:
    """
    Routing priority:
      1) If any row has GL code == '1200'           -> 'SHOP'
      2) Else if any row has WO present             -> 'WO'
      3) Else if any row has non-'shop' remarks AND GL != '1200' -> 'WO'
      4) Else -> '' (blank)
    """
    # 1) Any GL=1200? -> SHOP
    if df_group["Distribution_GL_Account"].astype(str).str.strip().apply(strip_dotzero).eq("1200").any():
        return "SHOP"

    # 2) Any WO present?
    if (df_group["__WO_OUT__"].astype(str).str.strip() != "").any():
        return "WO"

    # 3) Non-shop remarks (accept typos) with GL != 1200 -> WO
    remark_exceptions = {"shop", "shop stock", "shop stocl", "shop stockl"}
    for _, r in df_group.iterrows():
        gl_code = strip_dotzero(r.get("Distribution_GL_Account", ""))
        rmk = clean_str(r.get("Remarks", "")).lower()
        if rmk and (rmk not in remark_exceptions) and gl_code != "1200":
            return "WO"

    # 4) Default
    return ""

# Group by invoice number to emit H then Ds
grp = work_df.groupby("Invoice_Number", dropna=False, sort=False)

H_EXPECT = 24
D_EXPECT = 36

lines_written = 0
h_count = 0
d_count = 0

os.makedirs(os.path.dirname(out_txt_path), exist_ok=True)
with open(out_txt_path, "w", newline="", encoding="utf-8") as f:
    for inv_no, g in grp:
        rows = g.sort_index()
        r0 = rows.iloc[0]

        # Compute routing for the entire invoice group
        routing_code = compute_routing_for_group(rows)

        # -------- Build full H (24 fields) --------
        vendor_code = clean_str(r0.get("Vendor_Code", ""), 10)
        inv_number  = clean_str(r0.get("Invoice_Number", ""), 20)
        inv_type    = clean_str(r0.get("Invoice_Type", "I"), 1) or "I"
        inv_date    = fmt_date_mmddyyyy(r0.get("Invoice_Date", ""))
        gl_date     = fmt_date_mmddyyyy(r0.get("GL_Date", ""))
        batch_code  = clean_str(r0.get("Batch_Code", ""), 10)
        header_cc   = clean_str(r0.get("Liability_Cost_Center", ""), 10)

        # Prefer explicit header amount, else sum of detail amounts
        header_amt = clean_num(r0.get("Invoice_Amount", ""), 2, allow_negative=False)
        if header_amt == "":
            detail_total = 0.0
            for _, rx in rows.iterrows():
                v = clean_num(rx.get("Invoice_Amount", 0.0) or 0.0) or "0.00"
                detail_total += float(v)
            header_amt = f"{detail_total:.2f}"

        remarks = clean_str(r0.get("Remarks", ""), 30)
        image_filename = clean_str(r0.get("Image_Filename", ""))  # optional column

        header_full = [
            "H",             # 1  Record Type
            vendor_code,     # 2  Vendor
            inv_number,      # 3  Invoice #
            inv_type,        # 4  Type (I/C…)
            inv_date,        # 5  Invoice Date (MMDDYYYY)
            gl_date,         # 6  GL Date (MMDDYYYY)
            batch_code,      # 7  Batch
            "",              # 8  Subcontract #
            "",              # 9  Status (left blank per current file)
            "",              # 10 Bank Account
            "",              # 11 Check #
            "",              # 12 Cash GL
            "",              # 13 Check Date
            header_amt,      # 14 Invoice Amount
            "",              # 15 Retention
            "",              # 16 PO #
            remarks,         # 17 Remarks
            routing_code,    # 18 Routing
            "",              # 19 Card #
            image_filename,  # 20 Image filename
            "",              # 21 Liability Cost Center (left blank)
            "",              # 22 VAT Code
            "",              # 23 VAT Tax Amt
            "",              # 24 Discount Amt
        ]
        header_out = pad_or_trim_to(header_full, H_EXPECT)
        f.write(",".join(header_out) + "\n")
        h_count += 1
        lines_written += 1

        # -------- Details (36 fields) --------
        for _, rx in rows.iterrows():
            gl_code = strip_dotzero(rx.get("Distribution_GL_Account", ""))
            amt     = clean_num(rx.get("Invoice_Amount", ""), 2, allow_negative=True)

            job   = clean_str(rx.get("Job_Number", ""))
            phase = normalize_phase_code(rx.get("Phase_Code", ""))  # <<< normalized here
            ct    = clean_str(rx.get("Cost_Type", ""))
            if gl_code == "1200":
                job, phase, ct = "", "", ""  # shop expenses, no job distribution

            # format job like "123.4" -> "123.40"
            if re.fullmatch(r"\d+\.\d", job):
                try:
                    job = f"{float(job):.2f}"
                except Exception:
                    pass

            wo_out      = clean_str(rx.get("__WO_OUT__"))
            item_code   = clean_str(rx.get("__ITEM_CODE__"))
            comp_code   = "AA1"
            tax_code    = "NT"
            d_remarks   = clean_str(rx.get("Remarks", ""), 30)
            image_fn    = clean_str(rx.get("Detail_Image_Filename", "")) or image_filename
            cost_center = clean_str(rx.get("Expense_Cost_Center", ""), 10)

            # === ITEM DESC LOGIC ===
            # If WO present OR (remarks non-shop AND gl != 1200),
            # Item Desc = first character of item_code without leading '!'
            remark_exceptions = {"shop", "shop stock", "shop stocl", "shop stockl"}
            rmk_lower = d_remarks.lower()
            condition_for_desc = (wo_out != "") or (rmk_lower and (rmk_lower not in remark_exceptions) and gl_code != "1200")
            if condition_for_desc and item_code:
                item_desc = clean_str(item_code.lstrip("!")[:1])
            else:
                item_desc = ""

            detail_full = [
                "D",            # 1  Record Type
                vendor_code,    # 2  Vendor
                inv_number,     # 3  Invoice #
                inv_type,       # 4  Type
                gl_code,        # 5  GL Dist. Account
                amt,            # 6  Amount
                job,            # 7  Job
                phase,          # 8  Phase (normalized to 4 digits, except '1000')
                ct,             # 9  Cost Type
                d_remarks,      # 10 Remarks
                "",             # 11 Equip Code
                "",             # 12 Equip Cost Cat
                "",             # 13 UOM
                "",             # 14 Qty
                "",             # 15 Bill Item
                comp_code,      # 16 Company Code
                tax_code,       # 17 Tax Code
                wo_out,         # 18 Work Order
                item_code,      # 19 Item Code
                item_desc,      # 20 Item Desc (single char from Item Code when rules match)
                "",             # 21 Unit-Cost
                "",             # 22 WO Site Equip
                "",             # 23 WO Site Equip Comp
                "",             # 24 Service Contract
                "",             # 25 Unit Bill
                image_fn,       # 26 Image filename
                "",             # 27 PM Work Order
                "",             # 28 Cost Center
                "", "", "", "", "", "", "",  # 29-36 Labor/T+M or reserved
            ]
            detail_out = pad_or_trim_to(detail_full, D_EXPECT)
            f.write(",".join(detail_out) + "\n")
            d_count += 1
            lines_written += 1

print(f"SUCCESS: wrote {out_txt_path}")
print("=== COMMA STRUCTURE VALIDATION SUMMARY ===")
print(f"H lines written: {h_count} (each exactly {H_EXPECT} fields)")
print(f"D lines written: {d_count} (each exactly {D_EXPECT} fields)")
print(f"Total lines: {lines_written}")

import os

# List of folders to clear
folders_to_clear = [
    "data/cropped_images",
    "data/image_of_pos",
    "data/OCR_text_Test",
    "data/po_ocr_output",
    "data/processed",
    "data/raw_pdfs"
]

# Loop through each folder and delete files
for folder in folders_to_clear:
    if os.path.exists(folder):
        print(f"Clearing folder: {folder}")
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
            except Exception as e:
                print(f"Error deleting {file_path}: {e}")
    else:
        print(f"Folder does not exist: {folder}")

print("SUCCESS: All files cleared successfully.")

# Define processed_dir for completion flag
processed_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'processed')
os.makedirs(processed_dir, exist_ok=True)

# Create completion flag
completion_flag_path = os.path.join(processed_dir, 'processing_complete.flag')
with open(completion_flag_path, 'w') as f:
    f.write('complete')

print("COMPLETE: Processing pipeline completed successfully!")
